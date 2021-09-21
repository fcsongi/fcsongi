import anvil.server
import os
import subprocess
from anvil.tables import app_tables
import anvil.media
from datetime import datetime
import queue, threading
import pandas, json, requests, glob

anvil.server.connect("GFSQBSM43YRN5U7VSVNZQOUK-KOUTRITYHIAQNZVZ")

def worker():
    while True:
        item = q.get()
        os.system('mv ./queue/"' + item.getFileName() + '" ./projects/"' + item.getFileName() + '"')
        print('Working on ' + item.getFileName())

        row = app_tables.projects.get_by_id(item.getFileID())
        if row is not None:
            row["confirmed_at"] = datetime.now()
            row["status"] = "calculating"
        else:
            print("This project has been deleted")
            print("Cleaning up..")
            os.system('rm ./projects/"' + item.getFileName() + '"')
            q.task_done()
            continue

        try:
            deal_id = create_deal_in_pipedrive(item.getFileName(),item.get_users_pipedrive_id())

            files = glob.glob("./dealId/*.txt")
            for file in files:
                os.system('rm ' + file)

            with open("./dealId/" + str(deal_id) + ".txt","w") as f:
                f.close()

            row = app_tables.projects.get_by_id(item.getFileID())
            if row is not None:
                output = subprocess.run("node ./node_modules/webgme-engine/src/bin/run_plugin.js -b master -n CMBSD_master CreateProjectv2 GlobalDelivery", capture_output=True, shell=True, text=True)
                print(output.stdout)
            os.system('rm ./dealId/' + str(deal_id) + ".txt")

            row = app_tables.projects.get_by_id(item.getFileID())
            if row is not None:
                update_deal_in_pipedrive(item.getFileName(),deal_id)
                upload_file_to_pipedrive(item.getFileName(),deal_id)
                send_file_to_portal(item=item ,row=row ,deal_id=deal_id)
                row['calculated_at'] = datetime.now()
                row['status'] = "calculated"
            else:
                print("This project has been deleted")
                print("Cleaning up..")
                api_token = get_pipedrive_api_token()
                r = requests.put('https://combridge.pipedrive.com/api/v1/deals/{id}?api_token={api_token}'.format(id=deal_id, api_token = api_token), json={"status":"lost"})
                if r.status_code == 200:
                    print("Deal successfully removed from pipedrive")
                else:
                    print("Deal could not be remoced prom pipedrive")
                    raise Exception()
                os.system('rm ./projects/"' + item.getFileName() + '"')
                q.task_done()
                continue
            os.system('rm ./projects/"' + item.getFileName() + '"')
            print('Finished succesfully ' + item.getFileName())
        except Exception as e:
            print('Ran into an error. Saving file and errors..')
            row = app_tables.projects.get_by_id(item.getFileID())
            if row is not None:
                row['status'] = "internal error"
                print("Internal error label to portal")
            with open('./errors/"' + item.getFileName().replace(".xlsx", ".txt") + '"', "w") as f:
                f.write(output.stdout)
                f.write(output.stderr)
                f.write("Pipedrive")
                f.write(str(e) + '\n')
                f.close()
            os.system('mv ./projects/"' + item.getFileName() + '" ./errors/"' + item.getFileName() + '"')
        q.task_done()

class ExcelFile():
    def __init__(self,file,file_id,user_pipedrive_id):
        self.file = file
        self.file_id = file_id
        self.user_pipedrive_id = user_pipedrive_id

    def getFileName(self):
        return self.file

    def getFileID(self):
        return self.file_id

    def get_users_pipedrive_id(self):
        return self.user_pipedrive_id

@anvil.server.background_task
def create_projectfile(sites, project_name, vendor, user):
    data = []
    for i in range(len(sites)):
        row = {
            'Site ID 1' : "",
            'Site ID 2' : "",
            'Overlay required?': "",
            'Underlay required?': "",
            'Country code' : "",
            'City' : "",
            'ZIP code': "",
            'Address': "",
            'Series': vendor,
            'Order type': "",
            'Bandwidth category': "",
            'Feature set': "",
            'Device type': "",
            'Device quantity': "",
            'Site setup': "", 
            'Installation required?': "",
            "Uninstallation required?":"",
            'Maintenance required?': "",
            'On-site maintenance': "",
            'Service type': "",
            'Download/Upload ratio': "",
            'Number of public IPs': "",
            'Network termination unit required?': "",
            'Contract term': ""
        }
        try:
            if 'site_id1' in sites[i].keys() and sites[i]['site_id1'] != "":
                row['Site ID 1'] = sites[i]['site_id1']
            else:
                row['Site ID 1'] = i + 1 
            if 'site_id2' in sites[i].keys() and sites[i]['site_id2'] != "":
                row['Site ID 2'] = sites[i]['site_id2']           
            row['Country code'] = sites[i]['country_code']
            if 'city' in sites[i].keys():
                row['City'] = sites[i]['city']
            if 'zip_code' in sites[i].keys():
                row['ZIP code'] = sites[i]['zip_code']
            if 'address' in sites[i].keys():
                row['Address'] = sites[i]['address']
            if sites[i]['overlay_req'] == "yes":
                row['Overlay required?'] = sites[i]['overlay_req'] #
                row['Order type'] = sites[i]['order_type']
                if sites[i]['order_type'] == "Bandwidth based":
                    row['Bandwidth category'] = sites[i]["bandwidth_category"]
                    if vendor != "Silverpeak":
                        row['Feature set'] = sites[i]["feature_set"]
                row['Device type'] = sites[i]['device']
                row['Device quantity'] = sites[i]['device_quantity']
                row['Site setup'] = sites[i]['site_setup']
                if sites[i]['install_req'] == "yes" or sites[i]['install_req'] == "Yes":
                    row['Installation required?'] = "yes" #
                else:
                    row['Installation required?'] = "no"
                if sites[i]['uninstall_req'] == "yes" or sites[i]['uninstall_req'] == "Yes":
                    row['Uninstallation required?'] = "yes" #
                else:
                    row['Uninstallation required?'] = "no"
                if sites[i]['maint_req'] == "yes" or sites[i]['maint_req'] == "Yes":
                    row['Maintenance required?'] = "yes" # 
                else:
                    row['Maintenance required?'] = "no"
                row['On-site maintenance'] = sites[i]['sla_category']
                j = 1
                for accessory in sites[i]['accessories']:
                    if accessory['accessory_needed']:
                        row['Accessory ' + str(j)] = accessory['accessory_code']
                        row['Quantity ' + str(j)] = accessory['accessory_quantity']
                        j+=1
            else:
                row['Overlay required?'] = "no"
            
            if sites[i]["underlay_req"] == "yes":
                row['Underlay required?'] = sites[i]['underlay_req']
                row['Service type']  = sites[i]['service_type']
                row['Bandwidth category'] = sites[i]["bandwidth_category"]
                row['Download/Upload ratio'] = sites[i]['download_upload_ratio']
                row['Number of public IPs'] = sites[i]['number_of_public_ips']
                if sites[i]['ntu_req'] == "yes":
                    row['Network termination unit required?'] = "yes" #
                else:
                    row['Network termination unit required?'] = "no"
            else:
                row['Underlay required?'] = "no"
            
            row['Contract term'] = sites[i]['contract_term']
        except Exception as e:
            with open('./errors/"' + project_name + '"', "w") as f:
                f.write("There are some keys that are missing from the file:" + str(e) + '\n')
                f.close()
            return

        data.append(row)
    df = pandas.DataFrame(data)
    df.to_excel("./queue/" + project_name + ".xlsx",index=False)
    
    media_object = anvil.media.from_file('./queue/' + project_name + ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    organization = None
    if user['organization'] is not None:
        organization = user['organization'] 
    project_row = app_tables.projects.add_row(file=media_object,project_name=project_name,created_at=datetime.now(),user=user,status="waiting for confirmation",organization=organization)

    excelFile = ExcelFile(project_name + ".xlsx",project_row.get_id(),user['pipedrive_id'])
    q.put(excelFile)

def send_file_to_portal(item, row, deal_id):
    import openpyxl
    #Creating quote excel for customer ( to portal )
    workbook = openpyxl.load_workbook(filename="./outputs/" + str(item.getFileName()))
    sheets = workbook.sheetnames

    for s in sheets:
        if s != "Quote" and s != "NonEu delivery" and s != "BOM":
            del workbook[s]

    for i in range(4,workbook["Quote"].max_row):
        workbook["Quote"]["A" + str(i)] = deal_id
    
    for rows in workbook["Quote"].iter_rows(min_row=1, max_row=workbook["Quote"].max_row, min_col=1):
        for cell in rows:
            if cell.row == 3 and cell.column >= 16 and cell.comment:
                cell.comment.width = 250
                cell.comment.height = 200

    quote_filename = str(item.getFileName()).replace(".xlsx", " Quote.xlsx")

    workbook.save(filename="./quotes/" + quote_filename)  

    quote_file = anvil.media.from_file('./quotes/' + quote_filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    row['quote_file'] = quote_file
    row['pipedrive_id'] = deal_id
    print("QUTENAME:" + item.getFileName())
    os.system('rm ./quotes/"' + item.getFileName().replace(".xlsx"," Quote.xlsx") + '"')

def write_a_media_object(media_object,filename):
    with open("./queue/"+filename, 'wb+') as f:
        f.write(media_object.get_bytes())

def create_deal_in_pipedrive(quoteName, person_id):
    from pipedrive.client import Client

    domain = 'https://combridge.pipedrive.com/'

    client = Client(domain=domain)

    client.set_api_token('d51836dbab77e5e248a7a0192b46b6e50dbae38f')

    r = client.persons.get_person(person_id)
    org_id = None
    if r["data"]["org_id"] is not None:
        org_id = r["data"]["org_id"]["value"]
        
    malwin_label_id = "559"
    compliance_check_id = "0ba80eeef0fcc7bbc81633f435feb9ee7cfbe066"
    compliance_flag_id = "b490eac1db0a2d9eb0989578bebef37be0a9cf20"
    data = {
        'title': quoteName.replace(".xlsx", ""),
        'person_id' : person_id,
        'org_id' : org_id,
        'label' : malwin_label_id,
        'pipeline_id' : 37,
        compliance_check_id:"Go",
        compliance_flag_id:"Off"
    }

    response = client.deals.create_deal(data)
    print("Deal created with ID:" + str(response['data']['id']))
    deal_id = response['data']['id']
    return deal_id

def update_deal_in_pipedrive(quoteName, deal_id):
    from pipedrive.client import Client

    domain = 'https://combridge.pipedrive.com/'

    client = Client(domain=domain)

    client.set_api_token('d51836dbab77e5e248a7a0192b46b6e50dbae38f')
    capex_id = "ae5ce3d8a71dcf2002296874dfcb696d34a5d233"
    ebitda_id = "1757af99ba2ab55614430f3f2a6c49942b86fa7c"
    cash_result_id = "e406d72c45d811056adda7b7238afdfa2d5b65dc"

    excel_df = pandas.read_excel("./outputs/" + quoteName,engine='openpyxl',dtype=object, sheet_name="Summary")
    result = excel_df.to_json(orient="records")
    sites = json.loads(result)

    data = {
        'currency' : 'USD',
        'value' : sites[0]['Revenue'],
        capex_id : sites[0]['Capex'],
        capex_id + '_currency' : 'USD',
        ebitda_id : sites[0]['Ebitda'],
        ebitda_id + '_currency' : 'USD',
        cash_result_id : sites[0]['Cash result'],
        cash_result_id + '_currency' : 'USD'
    }

    response = client.deals.update_deal(deal_id=deal_id,data=data)
    print(str(deal_id) + " deal has been updated with prices.")

def upload_file_to_pipedrive(quoteName, deal_id):
    from pipedrive.client import Client

    domain = 'https://combridge.pipedrive.com/'

    client = Client(domain=domain)
    client.set_api_token('d51836dbab77e5e248a7a0192b46b6e50dbae38f')

    files = {"file": (quoteName.replace(".xlsx"," Calc.xlsx"),open('./outputs/' + quoteName, 'rb'), "xlsx")}
    response = client.files.add_file(data={'deal_id':deal_id}, files=files)

def get_pipedrive_api_token():
    return "d51836dbab77e5e248a7a0192b46b6e50dbae38f"

@anvil.server.background_task
def send_file_to_server(file, filename, file_id):
    
    write_a_media_object(file, filename)

    excelFile = ExcelFile(filename,file_id)
    q.put(excelFile)

q = queue.Queue()
threading.Thread(target=worker, daemon= True).start()
anvil.server.wait_forever()
