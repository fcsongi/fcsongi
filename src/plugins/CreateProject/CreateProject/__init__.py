"""
This is where the implementation of the plugin code goes.
The CreateProject-class is imported from both run_plugin.py and run_debug.py
"""
import json
import sys
import logging
import pandas
import openpyxl
import glob
import time
from datetime import date
from webgme_bindings import PluginBase
import supplierLogic

# Setup a logger
logger = logging.getLogger("CreateProject")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)  # By default it logs to stderr..
handler.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
logger.addHandler(handler)

apac_countries = None
latam_countries = None
projectName = ""

#A class to save the device_country installation parameters
class NodePair:
    def __init__(self,bundleNode,countryNode,siteNode):
        self.bundleNode = bundleNode
        self.countryNode = countryNode
        self.siteNode = siteNode
        self.bundleGroupNode = None
        self.countryGroupNode = None
        self.installCost = 0.0
        self.additionalInstallCost = 0.0
        self.bronzeCost = 0.0
        self.additionalbronzeCost = 0.0
        self.silverCost = 0.0
        self.additionalSilverCost = 0.0
        self.goldCost = 0.0
        self.additionalGoldCost = 0.0
        self.ro_dst_deliveryCost = 0.0
        self.us_ro_deliveryCost = 0.0

    def get_bundleNode(self):
        return self.bundleNode

    def get_countryNode(self):
        return self.countryNode

    def get_siteNode(self):
        return self.siteNode

    def get_bundleGroupNode(self):
        return self.bundleGroupNode
    
    def get_countryGroupNode(self):
        return self.countryGroupNode

    def get_installCost(self):
        return self.installCost

    def get_additionalInstallCost(self):
        return self.additionalInstallCost

    def get_bronzeCost(self):
        return self.bronzeCost

    def get_additionalBronzeCost(self):
        return self.additionalBronzeCost

    def get_silverCost(self):
        return self.silverCost

    def get_additionalSilverCost(self):
        return self.additionalSilverCost

    def get_goldCost(self):
        return self.goldCost

    def get_additionalGoldCost(self):
        return self.additionalGoldCost

    def get_Ro_Destination_deliveryCost(self):
        return self.ro_dst_deliveryCost
    
    def get_Us_Ro_deliveryCost(self):
        return self.us_ro_deliveryCost

    def set_bundleGroupNode(self, bundleGroupNode):
        self.bundleGroupNode = bundleGroupNode

    def set_countryGroupNode(self, countryGroupNode):
        self.countryGroupNode = countryGroupNode

    def set_installCost(self, cost):
        self.installCost = cost

    def set_additionalInstallCost(self, cost):
        self.additionalInstallCost = cost

    def set_bronzeCost(self, cost):
        self.bronzeCost = cost
    
    def set_additionalBronzeCost(self, cost):
        self.additionalBronzeCost = cost

    def set_silverCost(self, cost):
        self.silverCost = cost
    
    def set_additionalSilverCost(self, cost):
        self.additionalSilverCost = cost
    
    def set_goldCost(self, cost):
        self.goldCost = cost
    
    def set_additionalGoldCost(self, cost):
        self.additionalGoldCost = cost

    def set_Ro_Destination_deliveryCost(self, cost):
        self.ro_dst_deliveryCost = cost

    def set_Us_Ro_deliveryCost(self,cost):
        self.us_ro_deliveryCost = cost

def getPrices(self, nodePair):
    """
    This function sets the prices to the nodepair class
    Returns None
    
    TODO refactor to new function get_peer_nodes
    """

    core = self.core
    path = core.load_instances(self.META["UnitPrices"])[0]["nodePath"]
    node = core.load_by_path(self.root_node,path)
    children = core.load_children(node)
    if children:
        for child in children:
            if core.is_connection(child) and core.get_pointer_path(child,"src") == nodePair.get_bundleGroupNode()["nodePath"] and core.get_pointer_path(child,"dst") == nodePair.get_countryGroupNode()["nodePath"]:
                nodePair.set_installCost(core.get_attribute(child, "installCost"))
                nodePair.set_additionalInstallCost(core.get_attribute(child, "additionalInstallCost"))
                nodePair.set_bronzeCost(core.get_attribute(child, "bronzeCost"))
                nodePair.set_additionalBronzeCost(core.get_attribute(child, "additionalBronzeCost"))
                nodePair.set_silverCost(core.get_attribute(child,"silverCost"))
                nodePair.set_additionalSilverCost(core.get_attribute(child,"additionalSilverCost"))
                nodePair.set_goldCost(core.get_attribute(child,"goldCost"))
                nodePair.set_additionalGoldCost(core.get_attribute(child,"additionalGoldCost"))
                return
    else:
        print("There are no UnitPrices in the database")
            
            
def connect_vendor_bundle_to_site(self, siteNode, projectNode, position_item, site, vendor):
    """
    This function creates a copy of the bundle named site["Device type"] and connects it to the node
    Returns None

    TODO refactor: add error handling with return types
    """

    core = self.core
    vendorNodes = core.load_children(self.META["Vendors"])
    if vendorNodes:
        for vendorNode in vendorNodes:
            if core.get_attribute(vendorNode, "name") == vendor:
                bundleNodes = core.load_children(vendorNode)
                if bundleNodes:
                    for bundleNode in bundleNodes:
                        if core.get_attribute(bundleNode, "name") == site["Device type"]:
                            connection = core.create_child(projectNode, self.META["Bundle2Site"])
                            instance = core.create_child(projectNode, bundleNode)
                            position_item["x"] -= 200
                            core.set_registry(instance, "position", position_item)
                            core.set_pointer(connection, "src", instance)
                            core.set_pointer(connection, "dst", siteNode)
                            return instance
                    logger.info("There is no bundle named: " + site["Device type"])
                else:
                    logger.info("There are no bundles in " + core.get_attribute(vendorNode, "name") + " vendor")

        logger.info("There is no vendor named " + core.get_attribute(vendorNode, "name"))
    else:
        logger.info("There are no Vendors in the database")

def load_all_countries(self):
    """
    This function loads all the countries that webgme contain so we do not have to reload them multiple times
    Returns the list of country nodes if succeds
            None if countries are missing from database

    TODO refactor: The requested countries may be added directly to the site so in the future we might no need this function
    """
    core = self.core
    regionNodes = core.load_children(self.META["Countries"])
    countryNodes = []
    if regionNodes:
        for regionNode in regionNodes:
            if core.get_base_type(regionNode) == self.META["Region"]:
                countryNodes += core.load_children(regionNode)
        return countryNodes
    else:
        print("There are no regions in the database")

#returns all the attributes of the node as a dict
def get_attributes_of_node(core, node):
    """
    This function gathers all the attributes and values of a node
    Returns the attributes and their values as a dict
    """
    attributes = {}
    for attribute in core.get_attribute_names(node):
        attributes[attribute] = core.get_attribute(node,attribute)
    return attributes

def connect_country_to_site(self, siteNode, projectNode, position_item, site, countryNodes):
    """
    This function create an instance of the selected country node and connects it to the site
    Returns the instance
    """
    core = self.core
    for countryNode in countryNodes:
        if core.get_attribute(countryNode, "isoCode2") == site["Country code"]:
            connection = core.create_child(projectNode, self.META["Site2Country"])
            instance = core.create_child(projectNode, countryNode)
            position_item["x"] += 400
            core.set_registry(instance, "position", position_item)
            core.set_pointer(connection, "src", instance)
            core.set_pointer(connection, "dst", siteNode)
            return instance

def create_site_in_gme(self, projectNode, countryNodes, site, i):
    """
    This function creates a site in the given project node and sets it attributes
    Returns a nodepair of the site node, bundle node and country node
    """
    core = self.core
    siteNode = core.create_child(projectNode, self.META["Site"])
    position_item = core.get_registry(projectNode, "position")
    position_item["y"] = position_item["y"] + 50 * i
    core.set_registry(siteNode, "position", position_item)
    core.set_attribute(siteNode, "name", str(site["Site ID 1"]))
    core.set_attribute(siteNode, "siteId2", str(site["Site ID 2"]))
    core.set_attribute(siteNode, "city", site["City"])
    core.set_attribute(siteNode, "zipCode", str(site["ZIP code"]))
    core.set_attribute(siteNode, "address", site["Address"])
    core.set_attribute(siteNode, "siteSetup", site["Site setup"])
    core.set_attribute(siteNode, "deviceQuantity", int(site["Device quantity"]))
    core.set_attribute(siteNode, "installReq", site["Installation required?"])
    core.set_attribute(siteNode, "maintReq", site["Maintenance required?"])
    core.set_attribute(siteNode, "orderType", site["Order type"])
    if site["Overlay required?"] == "yes":
        core.set_attribute(siteNode, "overlayReq", True)
    if site["Underlay required?"] == "yes":
        core.set_attribute(siteNode, "underlayReq", True)
    if site["Order type"] == "Bandwidth based":
        core.set_attribute(siteNode, "bandwidth", site["Bandwidth category"])
        if "Feature set" in site.keys():
            core.set_attribute(siteNode, "featureSet", site["Feature set"])
    core.set_attribute(siteNode, "maintenanceType", site["On-site maintenance"])

    bundleNode = connect_vendor_bundle_to_site(self, siteNode, projectNode, position_item, site, site["Series"])
    countryNode = connect_country_to_site(self, siteNode, projectNode, position_item, site, countryNodes)
    nodePair = NodePair(bundleNode,countryNode,siteNode)
    return nodePair

def create_project_in_gme(self, active_node, projectName, contract_term, series):
    """
    This function creates a project node in the webgme and sets it attributes
    Returns the created project node
    """
    core = self.core
    projectMeta = self.META["Project"]
    projectNode = core.create_child(active_node, projectMeta)
    core.set_attribute(projectNode, "name", projectName)
    core.set_attribute(projectNode, "date", date.today().strftime("%d/%m/%Y"))
    core.set_attribute(projectNode, "contract term", int(contract_term))
    core.set_attribute(projectNode, "series", series)
    return projectNode

def get_group_nodes(self, nodePair):
    """
    This function sets the node pairs group node
    """
    core = self.core
    parent_node = core.get_base(nodePair.get_bundleNode())
    relative_path = list(core.is_member_of(parent_node).keys())[0]
    nodePair.set_bundleGroupNode(core.load_by_path(self.root_node, relative_path))

    parent_node = core.get_base(nodePair.get_countryNode())
    relative_path = list(core.is_member_of(parent_node).keys())[0]
    nodePair.set_countryGroupNode(core.load_by_path(self.root_node, relative_path))

#gets the delivery costs from Delivery methods node
def get_Ro_Destination_DeliveryCost(self, nodePair, deviceQuantity, series, accessories):
    """
    This function sets the the delivery cost of the package from Ro to the destination
    It sets directly the nodePair's attribute
    Returns None
    """
    core = self.core
    children = core.load_children(self.META["Delivery methods"])    
    child = children[0]
    children = core.load_children(child)
    for child in children:
        packageWeight = deviceQuantity * core.get_attribute(nodePair.get_bundleNode(),"weight")
        if accessories:
            packageWeight += 2
            packageWeight = round(packageWeight)
        if core.get_attribute(child,"name") == str(packageWeight) + "_" + str(core.get_attribute(nodePair.get_countryNode(),"DHL code")):
            nodePair.set_Ro_Destination_deliveryCost(core.get_attribute(child, "cost"))
            return

def get_Vendor_Ro_DeliveryCost(self, nodePair, deviceQuantity, series, accessories):
    """
    This function sets the the delivery cost of the package from the Vendor's country to Ro
    It sets directly the nodePair's attribute
    Returns None
    """
    core = self.core
    vendorCountryDHLcode = 0
    if series == "Silverpeak":
        vendorCountryDHLcode = 5 #US dhl code Silverpeak devices warehouse
    else:
        vendorCountryDHLcode = 2 #NL dhl code Juniper and Velocloud warehouse

    children = core.load_children(self.META["Delivery methods"])
    child = children[0]
    children = core.load_children(child)
    for child in children:
        packageWeight = deviceQuantity * core.get_attribute(nodePair.get_bundleNode(),"weight")
        if accessories:
            packageWeight += 2
            packageWeight = round(packageWeight)
        if core.get_attribute(child,"name") == str(packageWeight) + "_" + str(vendorCountryDHLcode):
            nodePair.set_Us_Ro_deliveryCost(core.get_attribute(child, "cost"))
            return

def create_BOM_row(siteID,code,description,quantity,discount,unit_list,contract_term=1):
    """
    This function creates a row so later we can add it to the BOM dataframe
    Returns the created row
    """
    global bom
    bom_row = {
        "Site ID":siteID,
        "Code":code,
        "Description": description,
        "Quantity": quantity * contract_term,
        "Discount": discount,
        "Unit list": unit_list,
        "Unit net": unit_list * (1 - discount)
    }
    bom_row["Total Due"] = bom_row["Unit net"] * quantity * contract_term
    return bom_row

def append_to_BOM_df(df,bom):
    """
    This function appends the given bom row to the BOM dataframe
    Returns the modified dataframe
    """
    for i in range(len(bom)):
        df = df.append(bom[i],ignore_index=True)
        if i == len(bom)-1:
            df = df.append({'Site ID':"",'Code':"",'Description':"",'Quantity':"",'Discount':"",'Unit list':"",'Unit net':"",'Total Due':""},ignore_index=True)
    return df

def get_silverpeak_costs(self, nodePair, site, accessories):
    """
    This function gathers the silverpeak bundle's costs
    Returns a dict with the costs
    """
    bundleNode = nodePair.get_bundleNode()
    core = self.core
    bundle = {
        "hardware" : 0,
        "software" : 0,
        "support" : 0,
        "license" : 0,
        "accessories": 0,
        "BOM":[]
    }
    
    bundle["hardware"] = float(core.get_attribute(bundleNode, "cost")) * (1 - float(core.get_attribute(bundleNode, "discount")))
    bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(bundleNode, "vendorCode"),description=core.get_attribute(bundleNode, "description"), quantity=site["Device quantity"], discount=core.get_attribute(bundleNode, "discount"),unit_list=core.get_attribute(bundleNode, "cost")))
    bundle["weight"] = core.get_attribute(bundleNode, "weight")
    parts = core.load_children(bundleNode)
    for part in parts:
        if core.get_attribute(part, "type") == "Support" and "1M" in core.get_attribute(part,"name"):
            core.set_registry(part, "color", "#00FF00")
            bundle["support"] = core.get_attribute(part,"cost") * (1 - core.get_attribute(part,"discount")) * site["Contract term"]
            bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"), description=core.get_attribute(part, "description"), quantity=site["Device quantity"], discount=core.get_attribute(part, "discount"), unit_list=core.get_attribute(part, "cost"), contract_term=site['Contract term']))
    for accessory in accessories:
        for part in parts:
            if core.get_attribute(part, "type") == "Accessory" and accessory['code'] == core.get_attribute(part, "vendorCode"):
                core.set_registry(part, "color", "#00FF00")
                bundle["accessories"] += core.get_attribute(part,"cost") * (1 - float(core.get_attribute(bundleNode, "discount"))) * accessory['quantity']
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"), description=core.get_attribute(part, "description"), quantity=accessory['quantity'], discount=core.get_attribute(bundleNode, "discount"), unit_list=core.get_attribute(part, "cost")))
    return bundle

def get_juniper_costs(self, nodePair, site, accessories):
    """
    This function gathers the Juniper bundle's costs
    Returns a dict with the costs
    """
    bundleNode = nodePair.get_bundleNode()
    core = self.core
    bundle = {
        "hardware" : 0,
        "software" : 0,
        "support" : 0,
        "license" : 0,
        "accessories": 0,
        "BOM" : []
    }

    supportDelay = "ND"
    if site["On-site maintenance"] == "Silver" or site["On-site maintenance"] == "Gold":
        supportDelay = "SD"

    featureSet = "STD"
    if "Feature set" in site.keys() and site["Feature set"] == "Extended":
        featureSet = "EXT"

    bundle["hardware"] = float(core.get_attribute(bundleNode, "cost")) * (1 - float(core.get_attribute(bundleNode, "discount")))
    bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(bundleNode, "vendorCode"),description=core.get_attribute(bundleNode, "description"), quantity=site["Device quantity"], discount=core.get_attribute(bundleNode, "discount"),unit_list=core.get_attribute(bundleNode, "cost")))
    bundle["weight"] = core.get_attribute(bundleNode, "weight")
    parts = core.load_children(bundleNode)
    contract_term_in_years = str(site["Contract term"]) + "M"
    for part in parts:
        if contract_term_in_years in core.get_attribute(part, "name") and featureSet in core.get_attribute(part, "name"):
            if core.get_attribute(part, "type") == "Support" and ("SUP" in core.get_attribute(part,"name") or supportDelay in core.get_attribute(part,"name")):
                core.set_registry(part, "color", "#00FF00")
                bundle["support"] += core.get_attribute(part, "cost") * (1 - core.get_attribute(part, "discount"))
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=site["Device quantity"], discount=core.get_attribute(part, "discount"),unit_list=core.get_attribute(part, "cost")))
            elif core.get_attribute(part, "type") == "Software":
                core.set_registry(part, "color", "#00FF00")
                bundle["software"] += core.get_attribute(part, "cost") * (1 - core.get_attribute(part, "discount"))
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=site["Device quantity"], discount=core.get_attribute(part, "discount"),unit_list=core.get_attribute(part, "cost")))
            elif core.get_attribute(part, "type") == "License":
                core.set_registry(part, "color", "#00FF00")
                bundle["license"] = core.get_attribute(part, "cost") * (1 - core.get_attribute(part, "discount"))
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=site["Device quantity"], discount=core.get_attribute(part, "discount"),unit_list=core.get_attribute(part, "cost")))
    bundle["hardware"] += bundle["software"]

    for accessory in accessories:
        for part in parts:
            if core.get_attribute(part, "type") == "Accessory" and accessory['code'] == core.get_attribute(part, "vendorCode"):
                core.set_registry(part, "color", "#00FF00")
                bundle["accessories"] += core.get_attribute(part, "cost") * (1 - core.get_attribute(bundleNode, "discount")) * accessory['quantity']
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"), description=core.get_attribute(part, "description"), quantity=accessory['quantity'], discount=core.get_attribute(bundleNode, "discount"), unit_list=core.get_attribute(part, "cost")))
    return bundle

def get_velocloud_costs(self, nodePair, site, accessories):
    """
    This function gathers the Velocloud bundle's costs
    Returns a dict with the costs
    """
    global apac_countries
    global latam_countries
    bundleNode = nodePair.get_bundleNode()
    countryNode = nodePair.get_countryNode()
    core = self.core
    bundle = {
        "hardware" : 0,
        "software" : 0,
        "support" : 0,
        "license" : 0,
        "accessories": 0,
        "BOM" : []
    }

    additionalDiscount = 0
    if site["Contract term"] >= 36:
        additionalDiscount = 0.03


    if "Feature set" in site.keys():
        featureSet = site["Feature set"].split("-")

    supportDelay = "NDD"
    if site["On-site maintenance"] == "Silver":
        supportDelay = "4H5"
    elif site["On-site maintenance"] == "Gold":
        supportDelay = "4H7"

    bundle["hardware"] = core.get_attribute(bundleNode, "cost") * (1 - core.get_attribute(bundleNode, "discount") - additionalDiscount)
    bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(bundleNode, "vendorCode"),description=core.get_attribute(bundleNode, "description"), quantity=site["Device quantity"], discount=core.get_attribute(bundleNode, "discount") + additionalDiscount,unit_list=core.get_attribute(bundleNode, "cost")))
    bundle["weight"] = core.get_attribute(bundleNode, "weight")
    if latam_countries is None and apac_countries is None:
        with open("./imports/LATAMcountries.csv", "r") as f:
            latam_countries = f.read().split(";")
        with open("./imports/APACcountries.csv", "r") as f:
            apac_countries = f.read().split(";")
    parts = core.load_children(bundleNode)
    contract_term = str(int(site["Contract term"])) + "P"

    bandwidth = 0
    if site["Order type"] == "Bandwidth based":
        if int(site["Bandwidth category"])//1000 == 0:
            if int(site["Bandwidth category"]) < 100:
                bandwidth = "0" + str(site["Bandwidth category"]) + "M"
            else:
                bandwidth = str(site["Bandwidth category"]) + "M"
        else:
            bandwidth = str(int(site["Bandwidth category"]) // 1000) + "G"

    for part in parts:
        if contract_term in core.get_attribute(part, "name"):
            if core.get_attribute(part, "type") == "Support" and supportDelay in core.get_attribute(part, "name"):
                core.set_registry(part, "color", "#00FF00")
                bundle["support"] = core.get_attribute(part, "cost") * (1 - core.get_attribute(bundleNode, "discount") - additionalDiscount) + core.get_attribute(part, "upgradeMargin") * bundle["hardware"]
                if site["Contract term"] >= 36:
                    bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=site["Device quantity"], discount=core.get_attribute(bundleNode, "discount") + additionalDiscount,unit_list=core.get_attribute(part, "cost")))
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code="VC-SUP-UPG-" + supportDelay + "-12P",description=core.get_attribute(part, "upgradeDescription"), quantity=site["Device quantity"], discount=core.get_attribute(bundleNode, "discount") + additionalDiscount,unit_list=core.get_attribute(part, "upgradeMargin") * core.get_attribute(bundleNode, "cost")))
            elif bandwidth != 0 :
                if site['Device quantity'] == 1:
                    quantity = 1
                else:
                    quantity = site['Device quantity'] // 2
                i = 0
                for i in range(len(featureSet)):
                    if featureSet[i] not in core.get_attribute(part,"name") :
                        break
                    elif i == len(featureSet) - 2:
                        if "G" not in featureSet[2] and "G" in core.get_attribute(part,"name"):
                            break
                        upgradeMargin = 0
                        if featureSet[-1] == "PROD":
                            upgradeCode = "VC-PROD-UPG-" + str(site["Contract term"]) + "P"
                            upgradeMargin = 0.05
                            upgradeDescription = "VMware SD-WAN support upgrade to Production, Subscription for " + str(int(site["Contract term"] / 12)) + " year, Prepaid"
                        elif featureSet[-1] == "PREM":
                            upgradeCode = "VC-PREM-UPG-" + str(site["Contract term"]) + "P"
                            upgradeMargin = 0.07
                            upgradeDescription = "VMware SD-WAN support upgrade to Premier, Subscription for " + str(int(site["Contract term"] / 12)) + " year, Prepaid"
                        
                        if core.get_attribute(part, "type") == "Software" and bandwidth in core.get_attribute(part, "name"):
                            core.set_registry(part, "color", "#00FF00")
                            bundle["software"] = core.get_attribute(part, "cost") * (1 - core.get_attribute(part, "discount") - additionalDiscount) + upgradeMargin * core.get_attribute(part, "cost") * (1 - core.get_attribute(part,"discount") - additionalDiscount)
                            bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=quantity, discount=core.get_attribute(part, "discount") + additionalDiscount,unit_list=core.get_attribute(part, "cost")))
                            if featureSet[-1] != "BAS":
                                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=upgradeCode,description=upgradeDescription, quantity=quantity, discount=core.get_attribute(part, "discount") + additionalDiscount, unit_list= upgradeMargin * core.get_attribute(part, "cost")))
                if core.get_attribute(part, "type") == "License" and featureSet[1] == "HO" and featureSet[2] == "HG" and bandwidth in core.get_attribute(part, "vendorCode") and ((core.get_attribute(countryNode, "name") in apac_countries and "APAC" in core.get_attribute(part,"name")) or (core.get_attribute(countryNode, "name") in latam_countries and "LATAM" in core.get_attribute(part,"name"))):
                    core.set_registry(part, "color", "#00FF00")
                    bundle["license"] = core.get_attribute(part, "cost") * (1 - core.get_attribute(bundleNode, "discount") - additionalDiscount)
                    bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"),description=core.get_attribute(part, "description"), quantity=quantity, discount=core.get_attribute(part, "discount") + additionalDiscount,unit_list=core.get_attribute(part, "cost")))
    for accessory in accessories:
        for part in parts:
            if core.get_attribute(part, "type") == "Accessory" and accessory['code'] == core.get_attribute(part, "vendorCode"):
                core.set_registry(part, "color", "#00FF00")
                bundle["accessories"] += core.get_attribute(part, "cost") * (1 - core.get_attribute(bundleNode, "discount") - additionalDiscount) * accessory['quantity']
                bundle["BOM"].append(create_BOM_row(siteID=str(site["Site ID 1"]) + str(site["Site ID 2"]),code=core.get_attribute(part, "vendorCode"), description=core.get_attribute(part, "description"), quantity=accessory['quantity'], discount=core.get_attribute(bundleNode, "discount") + additionalDiscount, unit_list=core.get_attribute(part, "cost")))
    return bundle

def get_install_uninstall_price(sites,i):
    """
    This function gathers the cost for the install and uninstall on the site
    """
    if sites[i]["Series"] != "Velocloud":
        if sites[i]["Site setup"] == "Standby":
            price = 600
        else:
            price = 600 + 150 * (sites[i]["Device quantity"] - 1)
    elif sites[i]["Series"] == "Velocloud":
        excel_df = pandas.read_excel("./imports/velocloudInstallPrices.xlsx",engine='openpyxl',dtype=object)
        result = excel_df.to_json(orient="records")
        installprices = json.loads(result)
        for prices in installprices:
            if sites[i]["Country code"] == prices["code"]:
                if sites[i]["Site setup"] == "Standby":
                    price = prices["base"]
                else:
                    price= prices["base"] + prices["additional"] * (sites[i]["Device quantity"] - 1)
    return price

def get_cell(sheet,name,i,max_col):
    """
    This function seraches a cell with the given column name and i (row)
    """
    for rows in sheet.iter_rows(min_row=i, max_row = i+1, min_col=1,max_col=max_col):
        for cell in rows:
            if name == cell.value:
                return cell 

def add_styles_to_excel(filename, vendor, MarginOnDelivery, overlay, underlay, underlayColumns):
    """
    This function adds the design to the excel we generated
    """
    workbook = openpyxl.load_workbook(filename="./outputs/" + filename + ".xlsx")
    if "Calc" in workbook.sheetnames:
        #Editing Calc sheet
        sheet = workbook['Calc']
        max_row = sheet.max_row
        max_col = sheet.max_column
        sheet["W" + str(max_row + 1)] = "=SUM(W2:W" + str(max_row) + ")"
        max_row += 1
        sheet["X" + str(max_row)] = "=SUM(X2:X" + str(max_row - 1) + ")"
        sheet["Y" + str(max_row)] = "=SUM(Y2:Y" + str(max_row - 1) + ")"
        sheet["AA" + str(max_row)] = "=SUM(AA2:AA" + str(max_row - 1) + ")"
        sheet["AB" + str(max_row)] = "=SUM(AB2:AB" + str(max_row - 1) + ")"
        sheet["AC" + str(max_row)] = "=SUM(AC2:AC" + str(max_row - 1) + ")"
        sheet["AD" + str(max_row)] = "=SUM(AD2:AD" + str(max_row - 1) + ")"
        sheet["AE" + str(max_row)] = "=SUM(AE2:AE" + str(max_row - 1) + ")"
        sheet["AF" + str(max_row)] = "=SUM(AF2:AF" + str(max_row - 1) + ")"
        sheet["AG" + str(max_row)] = "=SUM(AG2:AG" + str(max_row - 1) + ")"
        sheet["AH" + str(max_row)] = "=SUM(AH2:AH" + str(max_row - 1) + ")"
        if vendor == "Juniper":
            sheet["AI" + str(max_row)] = "=SUM(AI2:AI" + str(max_row - 1) + ")"
        elif vendor == "Velocloud":
            sheet["AI" + str(max_row)] = "=SUM(AI2:AI" + str(max_row - 1) + ")"
            sheet["AJ" + str(max_row)] = "=SUM(AJ2:AJ" + str(max_row - 1) + ")"
        sheet["W" + str(max_row + 2)] = "EU"
        max_row += 2
        sheet["X" + str(max_row)] = "=SUMIF(T2:T" + str(max_row - 3) + ",W" + str(max_row) + ",Y2:Y" + str(max_row - 3) + ")"
        sheet["W" + str(max_row + 1)] = "Not EU"
        max_row += 1
        sheet["X" + str(max_row)] = "=SUMIF(T2:T" + str(max_row - 4) + ",W" + str(max_row) + ",Y2:Y" + str(max_row - 4) + ")"

        accountant_format = u'_($* #,##0.00_);[Red]_($* (#,##0.00);_($* _0_0_);_(@'
        for rows in sheet.iter_rows(min_row=1, max_row=max_row+1, min_col=1):
            for cell in rows:
                if cell.row ==1:
                    cell.fill = openpyxl.styles.PatternFill(fgColor="808080", fill_type = "solid")
                    cell.font = openpyxl.styles.Font(color="FFFFFF")
                elif cell.column >= 23 and cell.column != 26:
                    cell.number_format = accountant_format
                else:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
                

        for i in range(1, max_col+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].bestFit = True
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].auto_size = True

        sheet.column_dimensions["G"].width = 50

    if "Quote" in workbook.sheetnames:
        #Editing Quote sheet

        sheet = workbook['Quote']
        sheet.insert_rows(idx=1, amount=2)
        max_row = sheet.max_row
        max_col = sheet.max_column
        sheet["F1"] = date.today().strftime("%d/%m/%Y")
        sheet["G1"] = "Quote expires in 60 days"
        sheet["G2"] = "This is a budgetary quote"
        sheet["A" + str(max_row + 1)] = "Subtotal"
        max_row += 1
        sheet.merge_cells("A" + str(max_row) + ":C" + str(max_row))

        if overlay:
            sheet["V1"] = "Combridge margin"
            sheet["V2"] = "Yearly finance margin"
            sheet["W1"] = "6%"
            sheet["W2"] = "4%"
            if vendor == "Juniper":
                sheet["AL2"] = "20%"
                sheet["AM2"] = "80%"
            elif vendor == "Velocloud":
                sheet["AM2"] = "20%"
                sheet["AN2"] = "80%"
            elif vendor == "Silverpeak":
                sheet["AK2"] = "20%"
                sheet["AL2"] = "80%"
            
            sheet["P3"].comment = openpyxl.comments.Comment('Final delivery costing component, door to door managed delivery service','Malwin')
            sheet["Q3"].comment = openpyxl.comments.Comment('Delivery cost for the pick-up of the devices from the warehouse of the vendor','Malwin')
            sheet["R3"].comment = openpyxl.comments.Comment('Physical uninstallation of the devices with remote support from our provisioning engineers','Malwin')
            sheet["S3"].comment = openpyxl.comments.Comment('Physical installation of the devices with remote support from our provisioning engineers','Malwin')
            sheet["T3"].comment = openpyxl.comments.Comment('Possible maintenance types:\nBronze - next business day\nSilver - 8x5x8\nGold - 24x7x4','Malwin')
            sheet["U3"].comment = openpyxl.comments.Comment('Yearly fee for the field engineer coverage, on site hand and eyes support','Malwin')
            sheet["V3"].comment = openpyxl.comments.Comment('Purchase cost of the equipment, derived from the vendor ','Malwin')
            sheet["W3"].comment = openpyxl.comments.Comment('Purchase cost of the accessories, derived from the vendor', 'Malwin')
            sheet["X3"].comment = openpyxl.comments.Comment('Purchase cost of the yearly vendor support, derived from the vendor ','Malwin')
            sheet["Y3"].comment = openpyxl.comments.Comment('Purchase cost of the devices lifted up with 6% Combridge margin','Malwin')
            sheet["Z3"].comment = openpyxl.comments.Comment('Purchase cost of the yearly  vendor support lifted up with 6% Combridge margin ','Malwin')
            sheet["AA3"].comment = openpyxl.comments.Comment('Net price of the equipment conform the number of equipment per each site, including Combridge margin','Malwin')
            sheet["AB3"].comment = openpyxl.comments.Comment('Net price of the accessories conform the number of accessories per each site, including Combridge margin','Malwin')
            sheet["AC3"].comment = openpyxl.comments.Comment('Net price of the yearly vendor support conform the number of equipment per each site, including Combridge margin','Malwin')
            sheet["AD3"].comment = openpyxl.comments.Comment('Yearly fee for the field engineer coverage, on site hand and eyes support','Malwin')
            sheet["AE3"].comment = openpyxl.comments.Comment('Uninstallation, Installation and delivery fees added together','Malwin')
            sheet["AF3"].comment = openpyxl.comments.Comment('Net price of the vendor support summed up for the contract term','Malwin')
            sheet["AG3"].comment = openpyxl.comments.Comment('Net price of the field service fee summed up for the contract term','Malwin')
            sheet["AH3"].comment = openpyxl.comments.Comment('Monthly cost element to cover the local licensing','Malwin')

            sheet["P" + str(max_row)] = "=SUM(P4:P" + str(max_row - 1) + ")"
            sheet["Q" + str(max_row)] = "=SUM(Q4:Q" + str(max_row - 1) + ")"
            sheet["R" + str(max_row)] = "=SUM(R4:R" + str(max_row - 1) + ")"
            sheet["S" + str(max_row)] = "=SUM(S4:S" + str(max_row - 1) + ")"
            sheet["U" + str(max_row)] = "=SUM(U4:U" + str(max_row - 1) + ")"
            sheet["V" + str(max_row)] = "=SUM(V4:V" + str(max_row - 1) + ")"
            sheet["W" + str(max_row)] = "=SUM(W4:W" + str(max_row - 1) + ")"
            sheet["X" + str(max_row)] = "=SUM(X4:X" + str(max_row - 1) + ")"
            sheet["Y" + str(max_row)] = "=SUM(Y4:Y" + str(max_row - 1) + ")"
            sheet["Z" + str(max_row)] = "=SUM(Z4:Z" + str(max_row - 1) + ")"
            sheet["AA" + str(max_row)] = "=SUM(AA4:AA" + str(max_row - 1) + ")"
            sheet["AB" + str(max_row)] = "=SUM(AB4:AB" + str(max_row - 1) + ")"
            sheet["AC" + str(max_row)] = "=SUM(AC4:AC" + str(max_row - 1) + ")"
            sheet["AD" + str(max_row)] = "=SUM(AD4:AD" + str(max_row - 1) + ")"
            sheet["AE" + str(max_row)] = "=SUM(AE4:AE" + str(max_row - 1) + ")"
            sheet["AF" + str(max_row)] = "=SUM(AF4:AF" + str(max_row - 1) + ")"
            sheet["AG" + str(max_row)] = "=SUM(AG4:AG" + str(max_row - 1) + ")"
            sheet["AH" + str(max_row)] = "=SUM(AH4:AH" + str(max_row - 1) + ")"

            if vendor == "Juniper":
                sheet["AI" + str(max_row)] = "=SUM(AI4:AI" + str(max_row - 1) + ")"
                sheet["AK" + str(max_row)] = "=SUM(AK4:AK" + str(max_row - 1) + ")"
                sheet["AL" + str(max_row)] = "=SUM(AL4:AL" + str(max_row - 1) + ")"
                sheet["AM" + str(max_row)] = "=SUM(AM4:AM" + str(max_row - 1) + ")"
                sheet["AN" + str(max_row)] = "=SUM(AN4:AN" + str(max_row - 1) + ")"

                sheet["AI3"].comment = openpyxl.comments.Comment('SD WAN Software licenses, lifted up with Combridge Margin','Malwin')
                sheet["AJ3"].comment = openpyxl.comments.Comment('Includes coordination and synchronisation  of the rollout from the procurement phase until the start of service on customer location, staging/pre-configuration activities if required','Malwin')
                sheet["AK3"].comment = openpyxl.comments.Comment('Value of the project management fee, calculated from the total value of costs over the contract term for each site separately','Malwin')
                sheet["AL3"].comment = openpyxl.comments.Comment("One time cost components summed up, includes delivery, installation and 20% of the hardware cost component  and the project management amount for each site. Can be the value on an eventual PO representing the amount payable after the successful installation  of the devices ",'Malwin')
                sheet["AM3"].comment = openpyxl.comments.Comment('Monthly recurring cost components summed up, includes 80% of the hardware value split over the contract term with the yearly financial up-lift, TC license additional fees, field support, vendor support eventual software cost components ','Malwin')
                sheet["AN3"].comment = openpyxl.comments.Comment('Total costs summed up over the contract term, the values should reflect the total amount on a PO in case of an eventual order ','Malwin')
                sheet.insert_cols(idx=openpyxl.utils.cell.column_index_from_string("AO"), amount=1)
                max_col += 1
                
            elif vendor == "Velocloud":
                sheet["AI" + str(max_row)] = "=SUM(AI4:AI" + str(max_row - 1) + ")"
                sheet["AJ" + str(max_row)] = "=SUM(AJ4:AJ" + str(max_row - 1) + ")"
                sheet["AL" + str(max_row)] = "=SUM(AL4:AL" + str(max_row - 1) + ")"
                sheet["AM" + str(max_row)] = "=SUM(AM4:AM" + str(max_row - 1) + ")"
                sheet["AN" + str(max_row)] = "=SUM(AN4:AN" + str(max_row - 1) + ")"
                sheet["AO" + str(max_row)] = "=SUM(AO4:AO" + str(max_row - 1) + ")"

                sheet["AI3"].comment = openpyxl.comments.Comment('Software ( bandwidth licenses etc ) related cost elements lifted up with Combridge Margin','Malwin')
                sheet["AJ3"].comment = openpyxl.comments.Comment('Software related additional cost element for specific regions around the world for gateway addon, LATAM/Asia lifted up with Combridge Margin','Malwin')
                sheet["AK3"].comment = openpyxl.comments.Comment('Includes coordination and synchronisation  of the rollout from the procurement phase until the start of service on customer location, staging/pre-configuration activities if required','Malwin')
                sheet["AL3"].comment = openpyxl.comments.Comment('Value of the project management fee, calculated from the total value of costs over the contract term for each site separately','Malwin')
                sheet["AM3"].comment = openpyxl.comments.Comment("One time cost components summed up, includes delivery, installation and 20% of the hardware cost component  and the project management amount for each site. Can be the value on an eventual PO representing the amount payable after the successful installation  of the devices ",'Malwin')
                sheet["AN3"].comment = openpyxl.comments.Comment('Monthly recurring cost components summed up, includes 80% of the hardware value split over the contract term with the yearly financial up-lift, TC license additional fees, field support, vendor support eventual software cost components ','Malwin')
                sheet["AO3"].comment = openpyxl.comments.Comment('Total costs summed up over the contract term, the values should reflect the total amount on a PO in case of an eventual order ','Malwin')
                sheet.insert_cols(idx=openpyxl.utils.cell.column_index_from_string("AP"), amount=1)
                max_col += 1

            elif vendor == "Silverpeak":
                sheet["AJ" + str(max_row)] = "=SUM(AJ4:AJ" + str(max_row - 1) + ")"
                sheet["AK" + str(max_row)] = "=SUM(AK4:AK" + str(max_row - 1) + ")"
                sheet["AL" + str(max_row)] = "=SUM(AL4:AL" + str(max_row - 1) + ")"
                sheet["AM" + str(max_row)] = "=SUM(AM4:AM" + str(max_row - 1) + ")"

                sheet["AI3"].comment = openpyxl.comments.Comment('Includes coordination and synchronisation  of the rollout from the procurement phase until the start of service on customer location, staging/pre-configuration activities if required','Malwin')
                sheet["AJ3"].comment = openpyxl.comments.Comment('Value of the project management fee, calculated from the total value of costs over the contract term for each site separately','Malwin')
                sheet["AK3"].comment = openpyxl.comments.Comment("One time cost components summed up, includes delivery, installation and 20% of the hardware cost component  and the project management amount for each site. Can be the value on an eventual PO representing the amount payable after the successful installation  of the devices ",'Malwin')
                sheet["AL3"].comment = openpyxl.comments.Comment('Monthly recurring cost components summed up, includes 80% of the hardware value split over the contract term with the yearly financial up-lift, TC license additional fees, field support, vendor support eventual software cost components ','Malwin')
                sheet["AM3"].comment = openpyxl.comments.Comment('Total costs summed up over the contract term, the values should reflect the total amount on a PO in case of an eventual order ','Malwin')
                sheet.insert_cols(idx=openpyxl.utils.cell.column_index_from_string("AN"), amount=1)
                max_col += 1

            for rows in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1):
                for cell in rows:
                    if cell.row <= 2:
                        if cell.column == 22 or cell.column == 23 or (cell.column >= 3 or cell.column <= 5) and cell.row == 1:
                            cell.font = openpyxl.styles.Font(bold=True)
                        else:
                            cell.font = openpyxl.styles.Font(size=16, bold=True)
                    elif cell.row == 3 and cell.column >= 16 and cell.comment:
                        cell.fill = openpyxl.styles.PatternFill(fgColor="ffcc99", fill_type = "solid")
                        cell.comment.width = 250
                        cell.comment.height = 200
                    elif cell.column >= 16 and sheet[str(openpyxl.utils.get_column_letter(cell.column)) + "3"].value is not None and "Quantity " not in sheet[str(openpyxl.utils.get_column_letter(cell.column)) + "3"].value:
                        cell.number_format = accountant_format
                    if cell.row > 3 and cell.row < max_row:
                        if cell.column == get_cell(sheet, "OTC", 3,max_col).column:
                            cell.value = "=" + str(get_cell(sheet, "Uninstallation + Installation + Delivery", 3,max_col).column_letter) + str(cell.row) + "+(" + str(get_cell(sheet, "Equipment total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet, "Accessories total price",3,max_col).column_letter) + str(cell.row) + ")*" \
                                + str(get_cell(sheet, "OTC",3,max_col).column_letter) + str(get_cell(sheet, "OTC",3,max_col).row-1) + "+" + str(get_cell(sheet,"Project Management",3,max_col).column_letter) + str(cell.row)
                        elif cell.column == get_cell(sheet, "MRC", 3,max_col).column:
                            if vendor == "Velocloud":
                                cell.value = "=" + str(get_cell(sheet, "TC MRC License price",3,max_col).column_letter) + str(cell.row) + "+((" + str(get_cell(sheet, "Equipment total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Accessories total price",3,max_col).column_letter) + str(cell.row) + ")*" \
                                    + str(get_cell(sheet,"MRC",3,max_col).column_letter) + str(get_cell(sheet,"MRC",3,max_col).row-1) + "*(1+" + str(openpyxl.utils.cell.get_column_letter(get_cell(sheet,"Yearly finance margin",2,max_col).column+1)) + "2)^(" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) \
                                    + str(cell.row) + "/12)+" + str(get_cell(sheet,"HW maintenance total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Total field service price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Software license price",3,max_col).column_letter) \
                                    + str(cell.row) + "+" + str(get_cell(sheet,"Region addon price",3,max_col).column_letter) + str(cell.row) +")/" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) + str(cell.row)
                            elif vendor == "Juniper":
                                cell.value = "=" + str(get_cell(sheet, "TC MRC License price",3,max_col).column_letter) + str(cell.row) + "+((" + str(get_cell(sheet, "Equipment total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Accessories total price",3,max_col).column_letter) + str(cell.row) + ")*" \
                                    + str(get_cell(sheet,"MRC",3,max_col).column_letter) + str(get_cell(sheet,"MRC",3,max_col).row-1) + "*(1+" + str(openpyxl.utils.cell.get_column_letter(get_cell(sheet,"Yearly finance margin",2,max_col).column+1)) + "2)^(" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) \
                                    + str(cell.row) + "/12)+" + str(get_cell(sheet,"HW maintenance total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Total field service price",3,max_col).column_letter) + str(cell.row)+ "+" + str(get_cell(sheet,"CSO license price",3,max_col).column_letter) \
                                    + str(cell.row) +")/" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) + str(cell.row)
                            elif vendor == "Silverpeak":
                                cell.value = "=" + str(get_cell(sheet, "TC MRC License price",3,max_col).column_letter) + str(cell.row) + "+((" + str(get_cell(sheet, "Equipment total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Accessories total price",3,max_col).column_letter) + str(cell.row) + ")*" \
                                    + str(get_cell(sheet,"MRC",3,max_col).column_letter) + str(get_cell(sheet,"MRC",3,max_col).row-1) + "*(1+" + str(openpyxl.utils.cell.get_column_letter(get_cell(sheet,"Yearly finance margin",2,max_col).column+1)) + "2)^(IF(" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) \
                                    + str(cell.row) + "/12<1,1,ROUND(" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) + str(cell.row) + "/12,0)))+" + str(get_cell(sheet,"HW maintenance total price",3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"Total field service price",3,max_col).column_letter) + str(cell.row) + ")/" + str(get_cell(sheet,"Contract term",3,max_col).column_letter) + str(cell.row)
                        elif cell.column == get_cell(sheet, "TCV", 3,max_col).column:
                            cell.value = "=" + str(get_cell(sheet, "OTC", 3,max_col).column_letter) + str(cell.row) + "+" + str(get_cell(sheet,"MRC",3,max_col).column_letter) + str(cell.row) + "*" + str(get_cell(sheet, "Contract term", 3,max_col).column_letter) + str(cell.row)
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

        if underlay:
            underlayFirstColumnIndex = 1
            for i in range(len(underlayColumns)):
                cell = get_cell(sheet,underlayColumns[i],3,max_col)
                if cell.value == "Service type" and overlay:
                    underlayFirstColumnIndex = cell.column
                    sheet.insert_cols(idx=underlayFirstColumnIndex, amount=1)
                    max_col += 1
                if cell.value == "Supplier OTC":
                    underlayFirstNumberColumnIndex = cell.column
                    break

            for rows in sheet.iter_rows(min_row=1,max_row=max_row,min_col=underlayFirstColumnIndex):
                i=0
                for cell in rows:
                    if cell.row == 3 and "Supplier OTC" == cell.value:
                        i+=1
                        textcell = sheet[cell.column_letter + str(cell.row-1)]
                        sheet.merge_cells(textcell.coordinate + ":" + openpyxl.utils.get_column_letter(textcell.column + 1) + str(textcell.row))
                        textcell.value = "Supplier" + str(i)
                        textcell.fill = openpyxl.styles.PatternFill(fgColor="ffcc99", fill_type = "solid")
                        textcell.font = openpyxl.styles.Font(size=16, bold=True)
                    if cell.column >= underlayFirstNumberColumnIndex and cell.row > 3:
                        cell.number_format = accountant_format
                        sheet[cell.column_letter + str(max_row)] = "=SUM(" + cell.column_letter + "4:" + cell.column_letter + str(max_row - 1) + ")"
                    cell.alignment = openpyxl.styles.Alignment(wrap_text = True, horizontal="center", vertical="center")

        for i in range(1, max_col+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].bestFit = True
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].auto_size = True

        sheet.column_dimensions["G"].width = 50
    
    if "Summary" in workbook.sheetnames:
        #Editing summary sheet
        sheet = workbook["Summary"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        for rows in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1):
            for cell in rows:
                if cell.value <= 1 and cell.value > 0:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE_00
                else:
                    cell.number_format = accountant_format

        for i in range(1, max_col+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].bestFit = True
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].auto_size = True

    if "BOM" in workbook.sheetnames:
        #Editing BOM sheet
        sheet = workbook["BOM"]
        sheet.insert_rows(idx=0, amount=2)
        max_row = sheet.max_row
        max_col = sheet.max_column
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 125
        sheet["E1"] = "Total"
        sheet["F1"] = "=SUM(F4:F" + str(max_row) + ")"
        sheet["G1"] = "=SUM(G4:G" + str(max_row) + ")"
        sheet["H1"] = "=SUM(H4:H" + str(max_row) + ")"

        for rows in sheet.iter_rows(min_row=1, max_row=max_row):
            for cell in rows:
                if cell.row != 3:
                    if cell.column == 5:
                        cell.number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE_00
                    elif cell.column > 5:
                        cell.number_format = accountant_format

        for i in range(6, max_col+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25

    if 'NonEu delivery' in workbook.sheetnames:
        #Editing NonEU delivery sheet
        sheet = workbook["NonEu delivery"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        for rows in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1):
            for cell in rows:
                if cell.column >= 3 and cell.column != 14:
                    cell.number_format = accountant_format
                    if cell.column >= 4:
                        cell.value *= (1 + MarginOnDelivery)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

        for i in range(1, max_col+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25

    workbook.save(filename="./outputs/" + filename + ".xlsx")

class CreateProject(PluginBase):
    def main(self):
        global projectName
        core = self.core
        root_node = self.root_node
        active_node = self.active_node
        start_time = time.time()
        MarginOnDelivery =core.get_attribute(self.META["CmbVariables"], "MarginOnDelivery")
        ResaleMargin = core.get_attribute(self.META["CmbVariables"], "ResaleMargin")
        YearlyFinanceMargin = core.get_attribute(self.META["CmbVariables"], "YearlyFinanceMargin")
        TCLicMargin = core.get_attribute(self.META["CmbVariables"], "TC_LicMargin")
        project_management = core.get_attribute(self.META["CmbVariables"], "ProjectManagement")

        #TODO import static file to webgne logic. Task in mentioned in the "Update countries logic that are not supported by Tecex".
        excel_df = pandas.read_excel("./imports/StaticNonEuDeliveries.xlsx",engine='openpyxl',dtype=object, na_filter=False)
        result = excel_df.to_json(orient="records")
        static_non_eu_deliveries = json.loads(result)

        #Storing country nodes so we do not have to load them all again
        countryNodes = load_all_countries(self)

        overlay = False
        underlay = False

        files = glob.glob("./projects/*.xlsx")
        for file in files:
            excel_df = pandas.read_excel(file,engine='openpyxl',dtype=object, na_filter=False)
            result = excel_df.to_json(orient="records")
            sites = json.loads(result)

            deal_id = glob.glob("./dealId/*.txt")[0].replace("./dealId/","").replace(".txt","")

            underlaySites = []
            numberOfSitesInCountry = {}

            #Stores the shipping device cost and weight so we can gather the prices accordingly
            countryShipments = {}

            #TODO judge if this list is needed or not
            nodePairs = []
            NonEu_IOR_delivery_BOM = []
            #TODO judge if this list should be added to the sites dictionary for data comprehension
            device_costs = []
            #This Dataframe stores the BOM sheet of the generated quote
            bundle_BOM_df = pandas.DataFrame({'Site ID':[],'Code':[],'Description':[],'Quantity':[],'Discount':[],'Unit list':[],'Unit net':[],'Total Due':[]})

            projectName = file.replace("./projects/","")
            projectName = projectName.replace(".xlsx","")

            writer = pandas.ExcelWriter("./outputs/" + projectName + ".xlsx", mode='w')

            projectNode = create_project_in_gme(self, active_node, projectName, sites[0]["Contract term"], sites[0]["Series"])

            #TODO add Velocloud SLA prices to the webgme logic
            if sites[0]["Series"] == "Velocloud":
                excel_df = pandas.read_excel("./imports/velocloudSLAprices.xlsx",engine='openpyxl',dtype=object)
                result = excel_df.to_json(orient="records")
                sla_costs_rows = json.loads(result)

            for i in range(len(sites)):
                sites[i]["Contract term"] = int(sites[i]["Contract term"])
                if sites[i]["Overlay required?"] == "yes":
                    overlay = True
                    nodePair = create_site_in_gme(self, projectNode, countryNodes, sites[i], i)
                    nodePairs.append(nodePair)
                    #stores the number of sites grouped by country code
                    if sites[i]["Country code"] in numberOfSitesInCountry:
                        numberOfSitesInCountry[sites[i]["Country code"]] += 1
                    else:
                        numberOfSitesInCountry[sites[i]["Country code"]] = 1

                    #Gathers the accessories to a list
                    accessories = []
                    for j in range(1, 6):
                        if "Accessory " + str(j) in sites[i].keys() and sites[i]["Accessory " + str(j)] != "":
                            accessories.append({'code':sites[i]["Accessory " + str(j)], "quantity": sites[i]["Quantity " + str(j)]})

                    #gathers data for the bundle costs
                    bundle = {}
                    if sites[i]["Series"] == "Silverpeak":
                        bundle = get_silverpeak_costs(self,nodePairs[i], sites[i], accessories)
                    elif sites[i]["Series"] == "Juniper":
                        bundle = get_juniper_costs(self,nodePairs[i], sites[i], accessories)
                    elif sites[i]["Series"] == "Velocloud":
                        bundle = get_velocloud_costs(self, nodePairs[i], sites[i], accessories)
                        if sites[i]["Maintenance required?"] == "yes":
                            if sites[i]["Series"] == "Velocloud":
                                sla_category = "NDD"
                                if sites[i]["On-site maintenance"] == "Silver":
                                    sla_category = "4H5"
                                elif sites[i]["On-site maintenance"] == "Gold":
                                    sla_category = "4H7"
                                additionalDiscount = 0
                                if sites[i]["Contract term"] >= 36:
                                    additionalDiscount = 0.03
                                for row in sla_costs_rows:
                                    if sla_category in row["code"] and str(sites[i]["Contract term"]) in row["code"]:
                                        bundle["BOM"].append(create_BOM_row(siteID=str(sites[i]["Site ID 1"]) + str(sites[i]["Site ID 2"]),code=row["code"], description=row["description"], quantity=sites[i]["Device quantity"], discount=core.get_attribute(nodePairs[i].get_bundleNode(), "discount") + additionalDiscount, unit_list=row["cost"]))
                        
                    device_costs.append(bundle)
                    bundle_BOM_df = append_to_BOM_df(df=bundle_BOM_df,bom=bundle["BOM"])

                    sites[i]["Volumetric weight"] = (device_costs[i]["weight"] * sites[i]["Device quantity"])
                    if accessories:
                        sites[i]["Volumetric weight"] += 2

                    #gathers data for NON EU delivery costs
                    if sites[i]["Country code"] in countryShipments:
                        countryShipments[sites[i]["Country code"]]["shipment value"] += bundle["hardware"] * sites[i]["Device quantity"] + bundle["accessories"]
                        countryShipments[sites[i]["Country code"]]["shipment weight"] += sites[i]["Volumetric weight"]
                    else:
                        shipmentPair = {}
                        shipmentPair["shipment value"] = bundle["hardware"] * core.get_attribute(nodePair.get_siteNode(),"deviceQuantity") + bundle["accessories"]
                        shipmentPair["shipment weight"] = sites[i]["Volumetric weight"]
                        countryShipments[sites[i]["Country code"]] = shipmentPair

            for i in range(len(sites)):
                print(json.dumps(sites[i], indent=4))
                sites[i].update({
                    "Quote ID":deal_id,
                    "countryNode":nodePairs[i].get_countryNode(),
                    "countryCode3":core.get_attribute(nodePairs[i].get_countryNode(), "isoCode3"),
                    "Country":"",
                    "Currency":"USD",
                    "DHL country code":"",
                    "EU/ Not EU":"",
                    "Licensing category":"",
                    "Number of sites":0,
                    "Unit hardware value":0,
                    "Accessories value":0,
                    "Net hardware value":0,
                    "Unit Vendor support":0,
                    "Net Vendor support":0,
                    "Vendor - RO delivery cost":0,
                    "RO - DST delivery cost":0,
                    "Install cost":0,
                    "Uninstall cost":0,
                    "On-site maintenance cost":0,
                    "TC MRC License cost":0,
                    "Software license cost":0,
                    "Region addon cost":0,
                    "CSO license cost":0,
                    "International freight (incl. import, compliance, taxes)":0,
                    "Freight":0,
                    "Uninstallation":0,
                    "Installation (1 or 2 device)":0,
                    "SLA category":"",
                    "SLA price / year":0,
                    "Equipment base cost":0,
                    "Accessories cost":0,
                    "HW maintenance monthly base cost":0,
                    "Equipment unit price":0,
                    "HW maintenance monthly unit price":0,
                    "Equipment total price":0,
                    "Accessories total price":0,
                    "HW maintenance monthly total price":0,
                    "Yearly Field service":0,
                    "Uninstallation + Installation + Delivery":0,
                    "HW maintenance total price":0,
                    "Total field service price":0,
                    "TC MRC License price":0,
                    "Project Management percentage":"",
                    "Project Management":0,
                    "OTC":0,
                    "MRC":0,
                    "TCV":0
                })
                if sites[i]["Overlay required?"] == "yes":
                    print(json.dumps(device_costs[i], indent=4))
                    countryAttributes = get_attributes_of_node(core,nodePairs[i].get_countryNode())
                    get_group_nodes(self, nodePairs[i])
                    getPrices(self, nodePairs[i])
                    get_Ro_Destination_DeliveryCost(self, nodePairs[i], sites[i]["Device quantity"], sites[i]["Series"], accessories)
                    get_Vendor_Ro_DeliveryCost(self, nodePairs[i], sites[i]["Device quantity"], sites[i]["Series"], accessories)

                    if sites[i]["Installation required?"] == "yes":
                        sites[i]["Install cost"] = nodePairs[i].get_installCost()
                    else:
                        sites[i]["Install cost"] = 0
                    if sites[i]["Uninstallation required?"] == "yes":
                        sites[i]["Uninstall cost"] = nodePairs[i].get_installCost()
                    else:
                        sites[i]["Uninstall cost"] = 0
                    #TODO additional install and maintenance costs figured out
                    sites[i]["Number of sites"] = numberOfSitesInCountry[sites[i]["Country code"]]
                    sites[i]["EU/ Not EU"] = countryAttributes["EU/Not EU"]
                    sites[i]["Country"] = countryAttributes["name"]
                    sites[i]["DHL country code"] = countryAttributes["DHL code"]
                    sites[i]["Licensing category"] = countryAttributes["deliveryCategory"]

                    sites[i]["Unit hardware value"] = device_costs[i]["hardware"]
                    sites[i]["Net hardware value"] = sites[i]["Device quantity"] * device_costs[i]["hardware"]
                    sites[i]["Accessories value"] = device_costs[i]["accessories"]
                    
                    sites[i]["Vendor - RO delivery cost"] = nodePairs[i].get_Us_Ro_deliveryCost()
                    if sites[i]["EU/ Not EU"] == "EU":    
                        sites[i]["RO - DST delivery cost"] = nodePairs[i].get_Ro_Destination_deliveryCost()
                    else:
                        for row in static_non_eu_deliveries:
                            if row["isoCode2"] == sites[i]["Country code"]:
                                if row["Malwin support"] == "Supported":
                                    if not "NonEu delivery" in countryShipments[sites[i]["Country code"]].keys():
                                        IOR_request_row = {
                                            "project_name":projectName,
                                            "quoteId":deal_id,
                                            "siteId":str(sites[i]["Site ID 1"]) + "-" + (sites[i]["Site ID 2"]),
                                            "countryNode":core.get_base(nodePairs[i].get_countryNode()),
                                            "countryCode2":sites[i]["Country code"],
                                            "countryCode3":sites[i]["countryCode3"],

                                            "shipment_weight":countryShipments[sites[i]["Country code"]]["shipment weight"],
                                            "number_of_sites":sites[i]["Number of sites"],
                                            "shipment_value":countryShipments[sites[i]["Country code"]]["shipment value"],

                                            "Ship To Country": sites[i]["Country"],
                                            "Chargable Weight" : countryShipments[sites[i]["Country code"]]["shipment weight"],
                                            "Shipment Value (USD)" : countryShipments[sites[i]["Country code"]]["shipment value"]
                                        }
                                        IOR_request_row = supplierLogic.initiateApiCallForSite(self=self,quotingsite=IOR_request_row,service="IOR provider",syslogger=logger)
                                        countryShipments[sites[i]["Country code"]]["NonEu delivery"] = IOR_request_row["Total Invoice Amount"]  / sites[i]["Number of sites"]
                                        NonEu_IOR_delivery_BOM.append(IOR_request_row)
                                    sites[i]["RO - DST delivery cost"] = countryShipments[sites[i]["Country code"]]["NonEu delivery"]
                                else:
                                    IOR_request_row = {
                                        "Ship To Country" : sites[i]["Country"],
                                        "Chargable Weight" : countryShipments[sites[i]["Country code"]]["shipment weight"],
                                        "Shipment Value (USD)" : countryShipments[sites[i]["Country code"]]["shipment value"],
                                        "IOR and Import Compliance Fee" : 0,
                                        "Admin Fee": 0,
                                        "Liability Cover Fee" : 0,
                                        "Tax and Duty" : 0,
                                        "Total - Customs Brokerage Cost" : 0,
                                        "Total - Clearance Costs" : 0,
                                        "Total - Handling Costs" : 0,
                                        "Total - License Cost" : 0,
                                        "Bank Fees" : 0,
                                        "Total Invoice Amount" : 0,
                                        "Notes" : "This is only an estimated cost of shipment fee, based on shipment value and weight."
                                    }
                                    for row in static_non_eu_deliveries:
                                        if row["isoCode2"] == sites[i]["Country code"]:
                                            IOR_request_row["Total Invoice Amount"] = row["Standard Lane Fee"] + IOR_request_row["Shipment Value (USD)"] * row["Taxes and duties percentage"]
                                            break
                                    sites[i]["RO - DST delivery cost"] = IOR_request_row["Total Invoice Amount"] / sites[i]["Number of sites"]
                                    NonEu_IOR_delivery_BOM.append(IOR_request_row)
                                break
                                
                    if sites[i]["Maintenance required?"] == "yes":
                        if sites[i]["Series"] == "Velocloud":
                            sla_category = "NDD"
                            if sites[i]["On-site maintenance"] == "Silver":
                                sla_category = "4H5"
                            elif sites[i]["On-site maintenance"] == "Gold":
                                sla_category = "4H7"
                            additionalDiscount = 0
                            if sites[i]["Contract term"] >= 36:
                                additionalDiscount = 0.03
                            for row in sla_costs_rows:
                                if sla_category in row["code"] and str(sites[i]["Contract term"]) in row["code"]:
                                    sites[i]["On-site maintenance cost"] = row["cost"] * (1 - core.get_attribute(nodePairs[i].get_bundleNode(), "discount") - additionalDiscount) * sites[i]["Device quantity"]
                        else:
                            if sites[i]["On-site maintenance"] == "Bronze":
                                sites[i]["On-site maintenance cost"] = nodePairs[i].get_bronzeCost()
                            elif sites[i]["On-site maintenance"] == "Silver":
                                sites[i]["On-site maintenance cost"] = nodePairs[i].get_silverCost()
                            elif sites[i]["On-site maintenance"] == "Gold":
                                sites[i]["On-site maintenance cost"] = nodePairs[i].get_goldCost()
                    else:
                        sites[i]["On-site maintenance cost"] = 0

                    sites[i]["Unit Vendor support"] = device_costs[i]["support"]
                    sites[i]["Net Vendor support"] = sites[i]["Device quantity"] * sites[i]["Unit Vendor support"]
                    if sites[i]["Series"] == "Juniper":
                        sites[i]["CSO license cost"] = device_costs[i]["license"] * sites[i]["Device quantity"]
                    elif sites[i]["Series"] == "Velocloud":
                        if int(sites[i]['Device quantity']) == 1:
                            sites[i]["Software license cost"] = device_costs[i]["software"]
                            sites[i]["Region addon cost"] = device_costs[i]["license"]
                        else:
                            if sites[i]["Site setup"] == "HA Cluster":
                                sites[i]["Software license cost"] = device_costs[i]["software"] * (int(sites[i]['Device quantity']))
                                sites[i]["Region addon cost"] = device_costs[i]["license"] * (int(sites[i]['Device quantity']))
                            else:
                                sites[i]["Software license cost"] = device_costs[i]["software"] * (int(sites[i]['Device quantity']) // 2)
                                sites[i]["Region addon cost"] = device_costs[i]["license"] * (int(sites[i]['Device quantity']) // 2)
                    sites[i]["TC MRC License cost"] = countryAttributes["MRC_license"]

                    #quote generating
                    sites[i]["International freight (incl. import, compliance, taxes)"] = sites[i]["RO - DST delivery cost"] * (1 + MarginOnDelivery)
                    sites[i]["Freight"] = sites[i]["Vendor - RO delivery cost"] * (1 + MarginOnDelivery)

                    if sites[i]["Installation required?"] == "yes":
                        sites[i]["Installation (1 or 2 device)"] = get_install_uninstall_price(sites,i) 
                    else:
                        sites[i]["Installation (1 or 2 device)"] = 0

                    if sites[i]["Uninstallation required?"] =="yes":
                        sites[i]["Uninstallation"] = get_install_uninstall_price(sites,i)
                    else:
                        sites[i]["Uninstallation"] = 0


                    sites[i]["SLA category"] = sites[i]["On-site maintenance"]

                    if sites[i]["Maintenance required?"] == "yes":
                        if sites[i]["Series"] == "Velocloud":
                            sites[i]["SLA price / year"] = sites[i]["On-site maintenance cost"] / (sites[i]["Contract term"] / 12)
                        else:
                            if sites[i]["SLA category"] == "Bronze":
                                sites[i]["SLA price / year"] = 100 + 25 * (sites[i]["Device quantity"] - 1)
                            elif sites[i]["SLA category"] == "Silver":
                                sites[i]["SLA price / year"] = 250 + 75 * (sites[i]["Device quantity"] - 1)
                            elif sites[i]["SLA category"] == "Gold":
                                sites[i]["SLA price / year"] = 350 + 100 * (sites[i]["Device quantity"] - 1)
                    else:
                        sites[i]["SLA price / year"] = 0
                        sites[i]["SLA category"] = ""
                        
                    sites[i]["Equipment base cost"] = sites[i]["Unit hardware value"]
                    sites[i]["Accessories cost"] = sites[i]["Accessories value"]
                    sites[i]["Equipment unit price"] = sites[i]["Equipment base cost"] * (1 + ResaleMargin)
                    sites[i]["Equipment total price"] = sites[i]["Equipment unit price"] * sites[i]["Device quantity"]
                    sites[i]["Accessories total price"] = sites[i]["Accessories cost"] * (1 + ResaleMargin)

                    if sites[i]["Series"] == "Silverpeak":
                        sites[i]["HW maintenance monthly base cost"] = sites[i]["Unit Vendor support"] / sites[i]["Contract term"]
                        sites[i]["HW maintenance monthly unit price"] = sites[i]["HW maintenance monthly base cost"] + sites[i]["HW maintenance monthly base cost"] * ResaleMargin
                        sites[i]["HW maintenance monthly total price"] = sites[i]["HW maintenance monthly unit price"] * sites[i]["Device quantity"]
                        sites[i]["HW maintenance total price"] = sites[i]["HW maintenance monthly total price"] * sites[i]["Contract term"]
                    else:
                        sites[i]["HW maintenance yearly base cost"] = sites[i]["Unit Vendor support"] / (sites[i]["Contract term"] / 12)
                        sites[i]["HW maintenance yearly unit price"] = sites[i]["HW maintenance yearly base cost"] + sites[i]["HW maintenance yearly base cost"] * ResaleMargin
                        sites[i]["HW maintenance yearly total price"] = sites[i]["HW maintenance yearly unit price"] * sites[i]["Device quantity"]
                        sites[i]["HW maintenance total price"] = sites[i]["HW maintenance yearly total price"] * (sites[i]["Contract term"] / 12)

                    if sites[i]["Series"] == "Velocloud":
                        sites[i]["Yearly Field service"] = sites[i]["SLA price / year"] * (1 + ResaleMargin)
                    else:
                        sites[i]["Yearly Field service"] = sites[i]["SLA price / year"]
                    sites[i]["Total field service price"] = sites[i]["Yearly Field service"] * (sites[i]["Contract term"] / 12)

                    sites[i]["Uninstallation + Installation + Delivery"] = sites[i]["International freight (incl. import, compliance, taxes)"] + sites[i]["Freight"] + sites[i]["Installation (1 or 2 device)"] + sites[i]["Uninstallation"]

                    sites[i]["TC MRC License price"] =  sites[i]["TC MRC License cost"] * (1 + TCLicMargin)
                    sites[i]["OTC"] = sites[i]["Uninstallation + Installation + Delivery"] + (sites[i]["Equipment total price"] + sites[i]["Accessories total price"]) * 0.2
                    if sites[i]["Series"] == "Velocloud":
                        sites[i]["Software license price"] = sites[i]["Software license cost"] * (1 + ResaleMargin)
                        sites[i]["Region addon price"] = sites[i]["Region addon cost"] * (1 + ResaleMargin)
                        sites[i]["MRC"] = sites[i]["TC MRC License price"] + ((sites[i]["Equipment total price"] + sites[i]["Accessories total price"]) * 0.8 * (1 + YearlyFinanceMargin) ** (sites[i]["Contract term"] /12) + sites[i]["HW maintenance total price"] + sites[i]["Total field service price"] + sites[i]["Software license price"] + sites[i]["Region addon price"]) / sites[i]["Contract term"]
                    elif sites[i]["Series"] == "Juniper":
                        sites[i]["CSO license price"] = sites[i]["CSO license cost"] * (1 + ResaleMargin)
                        sites[i]["MRC"] = sites[i]["TC MRC License price"] + ((sites[i]["Equipment total price"] + sites[i]["Accessories total price"]) * 0.8 * (1 + YearlyFinanceMargin) ** (sites[i]["Contract term"] /12) + sites[i]["HW maintenance total price"] + sites[i]["Total field service price"] + sites[i]["CSO license price"]) / sites[i]["Contract term"]
                    else:
                        if sites[i]["Contract term"] / 12 < 1:
                            contract_term = 1
                        else:
                            contract_term = round(sites[i]["Contract term"] / 12)
                        sites[i]["MRC"] = sites[i]["TC MRC License price"] + ((sites[i]["Equipment total price"] + sites[i]["Accessories total price"]) * 0.8 * (1 + YearlyFinanceMargin) ** contract_term + sites[i]["HW maintenance total price"] + sites[i]["Total field service price"]) / sites[i]["Contract term"]
                    sites[i]["TCV"] = sites[i]["OTC"] + sites[i]["MRC"] * sites[i]["Contract term"]
                    sites[i]["Project Management percentage"] = "6%"
                    sites[i]["Project Management"] = sites[i]["TCV"] * project_management
                    sites[i]["OTC"] += sites[i]["Project Management"]
                    sites[i]["TCV"] = sites[i]["OTC"] + sites[i]["MRC"] * sites[i]["Contract term"]
                    logger.info("Overlay row processed " + str(i))


                underlayColumns = []
                sites[i]["Supplier OTC cost"] = 0
                sites[i]["Supplier OTC"] = 0
                sites[i]["Supplier MRC cost"] = 0
                sites[i]["Supplier MRC"] = 0
                sites[i]["Supplier TCV"] = 0
                if sites[i]["Underlay required?"] == "yes":
                    underlay = True
                    sites[i]["Bandwidth category "] = sites[i]["Bandwidth category"]
                    underlayColumns = ["Service type", "Bandwidth category ","Download/Upload ratio", "Number of public IPs", "Network termination unit required?", "Supplier OTC", "Supplier MRC", "Supplier TCV"]
                    if sites[i]["Network termination unit required?"]:
                        sites[i]["Network termination unit required?"] = "yes"
                    else:
                        sites[i]["network termination unit required?"] = "no"
                    
                    underlaySite = {
                        "countryNode":core.get_base(nodePairs[i].get_countryNode()),
                        "rowId":i,
                        "project_name":projectName,
                        "quoteId":deal_id,
                        "siteId":str(sites[i]["Site ID 1"]) + "-" + str(sites[i]["Site ID 2"]),
                        "countryCode2":sites[i]["Country code"],
                        "countryCode3":sites[i]["countryCode3"],
                        "city":sites[i]["City"],
                        "address":sites[i]["Address"],
                        "bandwidth":sites[i]["Bandwidth category"],
                        "contract_term_month":sites[i]["Contract term"],
                        "contract_term_year":sites[i]["Contract term"] // 12 + 1 if sites[i]["Contract term"] % 12 != 0 else sites[i]["Contract term"] // 12
                    }
                    underlaySites.append(underlaySite)

            if underlaySites != []:
                for underlaySite in underlaySites:
                    underlaySite = supplierLogic.initiateApiCallForSite(self=self,quotingsite=underlaySite,service="Connectivity",syslogger=logger)
                    if "Supplier OTC cost" in underlaySite.keys() and "Supplier MRC cost" in underlaySite.keys():
                        sites[underlaySite["rowId"]]["Supplier OTC cost"] = underlaySite["Supplier OTC cost"]
                        sites[underlaySite["rowId"]]["Supplier OTC"] = underlaySite["Supplier OTC cost"] * 1.2
                        sites[underlaySite["rowId"]]["Supplier MRC cost"] = underlaySite["Supplier MRC cost"] 
                        sites[underlaySite["rowId"]]["Supplier MRC"] = underlaySite["Supplier MRC cost"] * 1.2
                        sites[underlaySite["rowId"]]["Supplier TCV"] = sites[underlaySite["rowId"]]["Supplier OTC"] + underlaySite["contract_term_month"] * sites[underlaySite["rowId"]]["Supplier MRC"]
            
            columns = []
            # writes data to excel
            if sites[0]["Series"] == "Velocloud":
                columns = ["Site ID 1", "Site ID 2","City","ZIP code","Address","Country code","Country","Order type","Bandwidth category","Feature set","Device type","Site setup","On-site maintenance","Installation required?","Uninstallation required?","Maintenance required?","Contract term",\
                    "Device quantity","DHL country code","EU/ Not EU","Licensing category","Number of sites","Unit hardware value","Accessories value","Net hardware value","Volumetric weight",\
                    "Unit Vendor support", "Net Vendor support","Vendor - RO delivery cost","RO - DST delivery cost","Install cost","Uninstall cost","On-site maintenance cost","TC MRC License cost", "Software license cost","Region addon cost"]
            elif sites[0]["Series"] == "Juniper":
                columns = ["Site ID 1", "Site ID 2","Country code","Country","City","ZIP code","Address","Order type","Bandwidth category","Feature set","Device type","Site setup","On-site maintenance","Installation required?","Uninstallation required?","Maintenance required?","Contract term",\
                    "Device quantity","DHL country code","EU/ Not EU","Licensing category","Number of sites","Unit hardware value","Accessories value","Net hardware value","Volumetric weight",\
                    "Unit Vendor support", "Net Vendor support","Vendor - RO delivery cost","RO - DST delivery cost","Install cost","Uninstall cost","On-site maintenance cost","TC MRC License cost", "CSO license cost"]
            else:
                columns = ["Site ID 1", "Site ID 2","Country code","Country","City","ZIP code","Address","Order type","Bandwidth category","Feature set","Device type","Site setup","On-site maintenance","Installation required?","Uninstallation required?","Maintenance required?","Contract term",\
                    "Device quantity","DHL country code","EU/ Not EU","Licensing category","Number of sites","Unit hardware value","Accessories value","Net hardware value","Volumetric weight",\
                    "Unit Vendor support", "Net Vendor support","Vendor - RO delivery cost","RO - DST delivery cost","Install cost","Uninstall cost","On-site maintenance cost","TC MRC License cost"]

            df = pandas.DataFrame(sites)
            df.to_excel(writer, sheet_name='Calc', index=False, columns=columns, float_format="%.2f")

            basecolumns = ["Quote ID", "Site ID 1", "Site ID 2","Country","ZIP code","City","Address","Contract term"]
            accessory_columns = []
            for j in range(1, 6):
                if "Accessory " + str(j) in sites[0].keys():
                    accessory_columns.append("Accessory " + str(j))
                    accessory_columns.append("Quantity " + str(j))
            
            columns = basecolumns
            if overlay:
                if sites[0]["Series"] == "Velocloud":
                    columns += ["Order type","Bandwidth category","Feature set","Device type","Device quantity","Site setup","Currency",\
                        "International freight (incl. import, compliance, taxes)","Freight","Uninstallation","Installation (1 or 2 device)","SLA category","SLA price / year",\
                        "Equipment base cost","Accessories cost","HW maintenance yearly base cost","Equipment unit price","HW maintenance yearly unit price",\
                        "Equipment total price","Accessories total price","HW maintenance yearly total price","Yearly Field service","Uninstallation + Installation + Delivery",\
                        "HW maintenance total price","Total field service price","TC MRC License price","Software license price","Region addon price","Project Management percentage","Project Management","OTC","MRC","TCV"] + accessory_columns
                elif sites[0]["Series"] == "Juniper":
                    columns += ["Order type","Bandwidth category","Feature set","Device type","Device quantity","Site setup","Currency",\
                        "International freight (incl. import, compliance, taxes)","Freight","Uninstallation","Installation (1 or 2 device)","SLA category","SLA price / year",\
                        "Equipment base cost","Accessories cost","HW maintenance yearly base cost","Equipment unit price","HW maintenance yearly unit price",\
                        "Equipment total price","Accessories total price","HW maintenance yearly total price","Yearly Field service","Uninstallation + Installation + Delivery",\
                        "HW maintenance total price","Total field service price","TC MRC License price","CSO license price","Project Management percentage","Project Management","OTC","MRC","TCV"] + accessory_columns
                else:
                    columns += ["Order type","Bandwidth category","Feature set","Device type","Device quantity","Site setup","Currency",\
                        "International freight (incl. import, compliance, taxes)","Freight","Uninstallation","Installation (1 or 2 device)","SLA category","SLA price / year",\
                        "Equipment base cost","Accessories cost","HW maintenance monthly base cost","Equipment unit price","HW maintenance monthly unit price",\
                        "Equipment total price","Accessories total price","HW maintenance monthly total price","Yearly Field service","Uninstallation + Installation + Delivery",\
                        "HW maintenance total price","Total field service price","TC MRC License price","Project Management percentage","Project Management","OTC","MRC","TCV"] + accessory_columns
            if underlay:
                columns+= underlayColumns

            df.to_excel(writer, sheet_name='Quote', index=False, columns=columns, float_format="%.2f")

            if NonEu_IOR_delivery_BOM:
                df = pandas.DataFrame(NonEu_IOR_delivery_BOM)
                columns = ["Ship To Country","Chargable Weight","Shipment Value (USD)","IOR and Import Compliance Fee","Admin Fee","Liability Cover Fee","Tax and Duty","Total - Customs Brokerage Cost","Total - Clearance Costs","Total - Handling Costs","Total - License Cost","Bank Fees","Total Invoice Amount","Notes"]
                df.to_excel(writer, sheet_name='NonEu delivery', columns=columns, index=False, float_format="%.2f")

            bundle_BOM_df.to_excel(writer, sheet_name='BOM', index=False, float_format="%.2f")

            summary = {
                "Revenue" : 0,
                "Costs" : 0,
                "Hardware cost" : 0,
                "Accessories cost": 0,
                "Vendor support" : 0,
                "Delivery" : 0,
                "Uninstallation":0,
                "Installation" : 0,
                "Field service" : 0,
                "MRC license" : 0,
                "Software license cost" : 0,
                "Region addon cost": 0,
                "CSO license cost": 0,
                "Underlay supplier cost": 0,
                "Cash result" : 0,
                "Capex" : 0,
                "Ebitda" : 0,
                "Margin" : 0
            }

            for i in range(len(sites)):
                summary["Revenue"] += sites[i]["TCV"] + sites[i]["Supplier TCV"]
                summary["Hardware cost"] += sites[i]["Net hardware value"]
                summary["Accessories cost"] += sites[i]["Accessories value"]
                summary["Vendor support"] += sites[i]["Net Vendor support"]
                summary["Delivery"] += sites[i]["Vendor - RO delivery cost"] + sites[i]["RO - DST delivery cost"]
                summary["Installation"] += sites[i]["Install cost"]
                summary["Uninstallation"] += sites[i]["Uninstall cost"]
                summary["Field service"] += sites[i]["On-site maintenance cost"]
                summary["MRC license"] += sites[i]["TC MRC License cost"] * sites[i]["Contract term"]
                if sites[i]["Series"] == "Velocloud":
                    summary["Software license cost"] += sites[i]["Software license cost"]
                    summary["Region addon cost"] += sites[i]["Region addon cost"]
                if sites[i]["Series"] == "Juniper":
                    summary["CSO license cost"] += sites[i]["CSO license cost"]
                if sites[i]["EU/ Not EU"] == "EU":
                    summary["Capex"] += sites[i]["Net hardware value"]
                summary["Underlay supplier cost"] += sites[i]["Supplier OTC cost"] + sites[i]["Supplier MRC cost"] * sites[i]["Contract term"]
            
            #Project management 6%
            summary["Costs"] = summary["Hardware cost"] + summary["Accessories cost"] + summary["Vendor support"] + summary["Delivery"] + summary["Installation"] + summary["Uninstallation"] + summary["Field service"] + summary["MRC license"] + summary["Software license cost"] + summary["Region addon cost"] + summary["CSO license cost"] + summary["Underlay supplier cost"]
            summary["Cash result"] = summary["Revenue"] - summary["Costs"]
            summary["Ebitda"] = summary["Cash result"] + summary["Capex"]
            summary["Margin"] = summary["Cash result"] / summary["Revenue"] if summary["Revenue"] else 0

            if sites[0]["Series"] == "Velocloud":
                columns = ["Revenue","Costs","Hardware cost","Accessories cost","Vendor support","Delivery","Installation","Uninstallation","Field service","MRC license","Software license cost","Region addon cost","Underlay supplier cost","Cash result","Capex","Ebitda","Margin"]
            elif sites[0]["Series"] == "Juniper":
                columns = ["Revenue","Costs","Hardware cost","Accessories cost","Vendor support","Delivery","Installation","Uninstallation","Field service","MRC license","CSO license cost","Underlay supplier cost","Cash result","Capex","Ebitda","Margin"]
            else:
                columns = ["Revenue","Costs","Hardware cost","Accessories cost","Vendor support","Delivery","Installation","Uninstallation","Field service","MRC license","Underlay supplier cost","Cash result","Capex","Ebitda","Margin"]


            temp = []
            temp.append(summary)
            df = pandas.DataFrame(temp)
            df.to_excel(writer, sheet_name='Summary', index=False, columns=columns, float_format="%.2f")
            writer.close()

            add_styles_to_excel(filename=projectName, vendor=sites[0]["Series"], MarginOnDelivery=MarginOnDelivery, overlay=overlay, underlay=underlay, underlayColumns=underlayColumns)

        logger.info("--- {} seconds ---".format((time.time() - start_time)))
                
        commit_info = self.util.save(root_node, self.commit_hash, "test", "Python plugin succesfully generated a quote")
        logger.info("committed :{0}".format(commit_info))
